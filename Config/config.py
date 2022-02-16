import openpyxl


class Test_Data:
    # print(cell.value)
    # sheet.cell(row=2, column=2).value = "Rahul"
    # print(sheet.cell(row=2, column=2).value)
    # print(sheet.max_row)
    # print(sheet.max_column)
    # print(sheet['A5'].value)

    def getTestData(self, test_name):
        try:
            data_dict = {}
            book = openpyxl.load_workbook("../Test_Data.xlsx")
            sheet = book.active
            for i in range(1, sheet.max_row + 1):  # to get rows
                if sheet.cell(row=i, column=1).value == test_name:

                    for j in range(2, sheet.max_column + 1):  # to get columns
                        data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            return data_dict
        except FileNotFoundError as e:
            print("Exception occured while performing file operation", e)
